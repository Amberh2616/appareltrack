from decimal import Decimal
import openpyxl
from io import BytesIO
from datetime import datetime
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django.db import transaction

from .models import SalesOrder, SalesOrderItem, ProductionOrder, MaterialRequirement
from .serializers import (
    SalesOrderSerializer,
    SalesOrderItemSerializer,
    ProductionOrderSerializer,
    ProductionOrderDetailSerializer,
    ProductionOrderCreateSerializer,
    MaterialRequirementSerializer,
    MaterialRequirementSimpleSerializer,
    CalculateMRPSerializer,
    GeneratePOSerializer,
)
from .services import MRPService


class SalesOrderViewSet(viewsets.ModelViewSet):
    queryset = SalesOrder.objects.all()
    serializer_class = SalesOrderSerializer


class SalesOrderItemViewSet(viewsets.ModelViewSet):
    queryset = SalesOrderItem.objects.all()
    serializer_class = SalesOrderItemSerializer


class ProductionOrderViewSet(viewsets.ModelViewSet):
    """
    ViewSet for ProductionOrder (大貨訂單)

    Endpoints:
    - GET /production-orders/ - List all
    - POST /production-orders/ - Create new
    - GET /production-orders/{id}/ - Get detail
    - PUT/PATCH /production-orders/{id}/ - Update
    - DELETE /production-orders/{id}/ - Delete
    - POST /production-orders/{id}/calculate-mrp/ - Calculate material requirements
    - POST /production-orders/{id}/generate-po/ - Generate purchase orders
    - GET /production-orders/{id}/requirements-summary/ - Get requirements summary
    - POST /production-orders/{id}/confirm/ - Confirm order
    """
    queryset = ProductionOrder.objects.select_related(
        'style_revision__style',
        'organization'
    ).prefetch_related(
        'material_requirements'
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'customer', 'style_revision']
    search_fields = ['po_number', 'order_number', 'customer']
    ordering_fields = ['created_at', 'order_date', 'delivery_date', 'total_quantity']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ProductionOrderDetailSerializer
        elif self.action == 'create':
            return ProductionOrderCreateSerializer
        return ProductionOrderSerializer

    def perform_create(self, serializer):
        # Auto-set organization from first available (demo mode)
        from apps.core.models import Organization
        org = Organization.objects.first()
        serializer.save(organization=org)

    @action(detail=True, methods=['post'])
    def calculate_mrp(self, request, pk=None):
        """
        Calculate material requirements for this production order.

        POST /api/v2/production-orders/{id}/calculate-mrp/
        Body: {
            "usage_scenario_id": "uuid" (optional),
            "default_wastage_pct": 5.00 (optional)
        }
        """
        order = self.get_object()

        serializer = CalculateMRPSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        usage_scenario = None
        if serializer.validated_data.get('usage_scenario_id'):
            from apps.costing.models import UsageScenario
            try:
                usage_scenario = UsageScenario.objects.get(
                    id=serializer.validated_data['usage_scenario_id']
                )
            except UsageScenario.DoesNotExist:
                return Response(
                    {'error': 'UsageScenario not found'},
                    status=status.HTTP_404_NOT_FOUND
                )

        default_wastage = serializer.validated_data.get(
            'default_wastage_pct',
            Decimal('5.00')
        )

        requirements = MRPService.calculate_requirements(
            production_order=order,
            usage_scenario=usage_scenario,
            default_wastage_pct=default_wastage
        )

        return Response({
            'message': f'Calculated {len(requirements)} material requirements',
            'requirements_count': len(requirements),
            'summary': MRPService.get_requirements_summary(order)
        })

    @action(detail=True, methods=['post'])
    def generate_po(self, request, pk=None):
        """
        Generate purchase orders from material requirements.

        POST /api/v2/production-orders/{id}/generate-po/
        Body: {
            "group_by_supplier": true (optional, default true)
        }
        """
        order = self.get_object()

        if not order.mrp_calculated:
            return Response(
                {'error': 'Please calculate MRP first'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = GeneratePOSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        group_by_supplier = serializer.validated_data.get('group_by_supplier', True)

        try:
            purchase_orders = MRPService.generate_purchase_orders(
                production_order=order,
                group_by_supplier=group_by_supplier
            )

            if not purchase_orders:
                return Response({
                    'message': 'No purchase orders generated (all requirements already ordered or no items to order)',
                    'purchase_orders': []
                })

            return Response({
                'message': f'Generated {len(purchase_orders)} purchase order(s)',
                'purchase_orders': [
                    {
                        'id': str(po.id),
                        'po_number': po.po_number,
                        'supplier': po.supplier.name,
                        'total_amount': float(po.total_amount),
                        'lines_count': po.lines.count()
                    }
                    for po in purchase_orders
                ]
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def requirements_summary(self, request, pk=None):
        """
        Get summary of material requirements.

        GET /api/v2/production-orders/{id}/requirements-summary/
        """
        order = self.get_object()
        summary = MRPService.get_requirements_summary(order)
        return Response(summary)

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """
        Confirm the production order.

        POST /api/v2/production-orders/{id}/confirm/
        """
        order = self.get_object()

        if order.status != 'draft':
            return Response(
                {'error': 'Can only confirm orders in draft status'},
                status=status.HTTP_400_BAD_REQUEST
            )

        order.status = 'confirmed'
        order.save(update_fields=['status'])

        return Response({
            'status': 'confirmed',
            'message': 'Production order confirmed'
        })

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get production order statistics.

        GET /api/v2/production-orders/stats/
        """
        qs = self.get_queryset()

        from django.db.models import Sum, Count

        stats = {
            'total': qs.count(),
            'by_status': {},
            'total_quantity': qs.aggregate(
                total=Sum('total_quantity')
            )['total'] or 0,
            'total_amount': float(qs.aggregate(
                total=Sum('total_amount')
            )['total'] or 0),
        }

        # Count by status
        status_counts = qs.values('status').annotate(count=Count('id'))
        for item in status_counts:
            stats['by_status'][item['status']] = item['count']

        return Response(stats)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        """
        Import production orders from Excel file.

        POST /api/v2/production-orders/import-excel/
        Content-Type: multipart/form-data
        Body: file=<excel_file>

        Excel format:
        | PO Number | Customer | Style Number | Color | Total Qty | XS | S | M | L | XL | XXL | Unit Price | Currency | Order Date | Delivery Date | Notes |
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'No file uploaded'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate file extension
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Read Excel file
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            # Expected column mapping
            col_map = {
                'PO Number': 'po_number',
                'Customer': 'customer',
                'Style Number': 'style_number',
                'Color': 'color',
                'Total Qty': 'total_qty',
                'XS': 'xs',
                'S': 's',
                'M': 'm',
                'L': 'l',
                'XL': 'xl',
                'XXL': 'xxl',
                'Unit Price': 'unit_price',
                'Currency': 'currency',
                'Order Date': 'order_date',
                'Delivery Date': 'delivery_date',
                'Notes': 'notes',
            }

            # Find column indices
            col_indices = {}
            for idx, header in enumerate(headers):
                if header in col_map:
                    col_indices[col_map[header]] = idx

            # Validate required columns
            required = ['po_number', 'customer', 'style_number']
            missing = [r for r in required if r not in col_indices]
            if missing:
                return Response(
                    {'error': f'Missing required columns: {missing}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get organization
            from apps.core.models import Organization
            org = Organization.objects.first()

            # Process rows
            created_orders = []
            errors = []
            row_idx = 1

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                try:
                    with transaction.atomic():
                        # Extract data
                        po_number = row[col_indices['po_number']] if 'po_number' in col_indices else None
                        customer = row[col_indices['customer']] if 'customer' in col_indices else None
                        style_number = row[col_indices['style_number']] if 'style_number' in col_indices else None

                        if not po_number or not customer or not style_number:
                            errors.append({
                                'row': row_idx,
                                'error': 'Missing required fields (PO Number, Customer, or Style Number)'
                            })
                            continue

                        # Find style revision
                        from apps.styles.models import Style, StyleRevision
                        style = Style.objects.filter(style_number__iexact=style_number).first()
                        if not style:
                            errors.append({
                                'row': row_idx,
                                'error': f'Style "{style_number}" not found in system'
                            })
                            continue

                        revision = style.current_revision
                        if not revision:
                            revision = StyleRevision.objects.filter(style=style).order_by('-created_at').first()

                        if not revision:
                            errors.append({
                                'row': row_idx,
                                'error': f'No revision found for style "{style_number}"'
                            })
                            continue

                        # Build size breakdown
                        size_breakdown = {}
                        sizes = ['xs', 's', 'm', 'l', 'xl', 'xxl']
                        for size in sizes:
                            if size in col_indices:
                                val = row[col_indices[size]]
                                if val and isinstance(val, (int, float)) and val > 0:
                                    size_breakdown[size.upper()] = int(val)

                        # Calculate total quantity
                        total_qty = row[col_indices.get('total_qty')] if 'total_qty' in col_indices else None
                        if not total_qty or total_qty == 0:
                            total_qty = sum(size_breakdown.values())

                        if total_qty == 0:
                            errors.append({
                                'row': row_idx,
                                'error': 'Total quantity is 0'
                            })
                            continue

                        # Parse dates
                        order_date = None
                        delivery_date = None

                        if 'order_date' in col_indices:
                            od = row[col_indices['order_date']]
                            if od:
                                if isinstance(od, datetime):
                                    order_date = od.date()
                                elif isinstance(od, str):
                                    try:
                                        order_date = datetime.strptime(od, '%Y-%m-%d').date()
                                    except ValueError:
                                        pass

                        if 'delivery_date' in col_indices:
                            dd = row[col_indices['delivery_date']]
                            if dd:
                                if isinstance(dd, datetime):
                                    delivery_date = dd.date()
                                elif isinstance(dd, str):
                                    try:
                                        delivery_date = datetime.strptime(dd, '%Y-%m-%d').date()
                                    except ValueError:
                                        pass

                        # Get unit price
                        unit_price = row[col_indices.get('unit_price')] if 'unit_price' in col_indices else Decimal('0')
                        if unit_price is None:
                            unit_price = Decimal('0')
                        unit_price = Decimal(str(unit_price))

                        # Get currency
                        currency = row[col_indices.get('currency')] if 'currency' in col_indices else 'USD'
                        if not currency:
                            currency = 'USD'

                        # Get notes
                        notes = row[col_indices.get('notes')] if 'notes' in col_indices else ''
                        color = row[col_indices.get('color')] if 'color' in col_indices else ''

                        if color:
                            notes = f"Color: {color}\n{notes}" if notes else f"Color: {color}"

                        # Generate order number
                        from apps.samples.services.auto_generation import get_next_sequence
                        seq = get_next_sequence('ORD')
                        order_number = f"ORD-{datetime.now().strftime('%y%m')}-{seq:06d}"

                        # Check for duplicate PO number
                        if ProductionOrder.objects.filter(po_number=po_number).exists():
                            errors.append({
                                'row': row_idx,
                                'error': f'PO Number "{po_number}" already exists'
                            })
                            continue

                        # Create production order
                        order = ProductionOrder.objects.create(
                            organization=org,
                            po_number=po_number,
                            order_number=order_number,
                            customer=customer,
                            style_revision=revision,
                            total_quantity=int(total_qty),
                            size_breakdown=size_breakdown,
                            unit_price=unit_price,
                            total_amount=unit_price * int(total_qty),
                            currency=currency,
                            order_date=order_date,
                            delivery_date=delivery_date,
                            notes=notes or '',
                            status='draft'
                        )

                        created_orders.append({
                            'row': row_idx,
                            'id': str(order.id),
                            'po_number': order.po_number,
                            'order_number': order.order_number,
                            'customer': order.customer,
                            'style_number': style_number,
                            'total_quantity': order.total_quantity,
                            'unit_price': float(order.unit_price),
                            'total_amount': float(order.total_amount),
                        })

                except Exception as e:
                    errors.append({
                        'row': row_idx,
                        'error': str(e)
                    })

            return Response({
                'success': True,
                'message': f'Imported {len(created_orders)} production order(s)',
                'created': created_orders,
                'errors': errors,
                'total_rows_processed': row_idx - 1
            })

        except Exception as e:
            return Response(
                {'error': f'Failed to process Excel file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MaterialRequirementViewSet(viewsets.ModelViewSet):
    """
    ViewSet for MaterialRequirement (物料需求)

    Read-only for most operations, as requirements are calculated by MRP service.
    """
    queryset = MaterialRequirement.objects.select_related(
        'production_order',
        'bom_item',
        'purchase_order_line'
    ).all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['production_order', 'status', 'category']
    search_fields = ['material_name', 'material_name_zh', 'supplier']
    ordering_fields = ['category', 'material_name', 'total_requirement']
    ordering = ['category', 'material_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return MaterialRequirementSimpleSerializer
        return MaterialRequirementSerializer

    @action(detail=True, methods=['patch'])
    def update_stock(self, request, pk=None):
        """
        Update current stock for a material requirement.

        PATCH /api/v2/material-requirements/{id}/update-stock/
        Body: {"current_stock": 100.00}
        """
        requirement = self.get_object()

        current_stock = request.data.get('current_stock')
        if current_stock is None:
            return Response(
                {'error': 'current_stock is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            requirement.current_stock = Decimal(str(current_stock))
            requirement.calculate_requirements()
            requirement.save()

            return Response({
                'id': str(requirement.id),
                'current_stock': float(requirement.current_stock),
                'order_quantity_needed': float(requirement.order_quantity_needed),
                'message': 'Stock updated'
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def review(self, request, pk=None):
        """
        Review and confirm a single material requirement.

        POST /api/v2/material-requirements/{id}/review/
        Body: {
            "quantity": "8610.0000" (optional, adjusted quantity),
            "unit_price": "2.50" (optional, confirmed unit price),
            "notes": "備註" (optional)
        }
        """
        from django.utils import timezone

        requirement = self.get_object()

        if requirement.is_reviewed:
            return Response(
                {'error': 'This requirement has already been reviewed'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update fields if provided
        if 'quantity' in request.data:
            requirement.reviewed_quantity = Decimal(str(request.data['quantity']))
        else:
            requirement.reviewed_quantity = requirement.order_quantity_needed

        if 'unit_price' in request.data:
            requirement.reviewed_unit_price = Decimal(str(request.data['unit_price']))
        elif requirement.bom_item and requirement.bom_item.unit_price:
            requirement.reviewed_unit_price = requirement.bom_item.unit_price
        else:
            requirement.reviewed_unit_price = Decimal('0')

        if 'notes' in request.data:
            requirement.review_notes = request.data['notes']

        if 'required_date' in request.data:
            from datetime import datetime
            req_date = request.data['required_date']
            if isinstance(req_date, str):
                requirement.required_date = datetime.strptime(req_date, '%Y-%m-%d').date()
            else:
                requirement.required_date = req_date
        elif not requirement.required_date:
            # Default to production order delivery date
            requirement.required_date = requirement.production_order.delivery_date

        if 'expected_delivery' in request.data:
            from datetime import datetime
            exp_date = request.data['expected_delivery']
            if isinstance(exp_date, str):
                requirement.expected_delivery = datetime.strptime(exp_date, '%Y-%m-%d').date()
            else:
                requirement.expected_delivery = exp_date

        requirement.is_reviewed = True
        requirement.reviewed_at = timezone.now()
        requirement.save()

        return Response({
            'id': str(requirement.id),
            'is_reviewed': True,
            'reviewed_quantity': str(requirement.reviewed_quantity),
            'reviewed_unit_price': str(requirement.reviewed_unit_price),
            'message': '物料需求已審核確認'
        })

    @action(detail=True, methods=['post'])
    def unreview(self, request, pk=None):
        """
        Unreview a material requirement for re-editing.

        POST /api/v2/material-requirements/{id}/unreview/
        """
        requirement = self.get_object()

        if requirement.status == 'ordered':
            return Response(
                {'error': 'Cannot unreview: PO already generated'},
                status=status.HTTP_400_BAD_REQUEST
            )

        requirement.is_reviewed = False
        requirement.reviewed_at = None
        requirement.save()

        return Response({
            'id': str(requirement.id),
            'is_reviewed': False,
            'message': '已取消審核，可重新編輯'
        })

    @action(detail=True, methods=['post'])
    def generate_po(self, request, pk=None):
        """
        Generate a single PurchaseOrder for this reviewed material requirement.

        POST /api/v2/material-requirements/{id}/generate-po/
        """
        from django.utils import timezone
        from apps.procurement.models import PurchaseOrder, POLine, Supplier, Material
        from apps.core.models import Organization

        requirement = self.get_object()

        # Validation
        if not requirement.is_reviewed:
            return Response(
                {'error': '請先審核此物料需求'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if requirement.status == 'ordered':
            return Response(
                {'error': '採購單已生成'},
                status=status.HTTP_400_BAD_REQUEST
            )

        order_qty = requirement.reviewed_quantity or requirement.order_quantity_needed
        if order_qty <= 0:
            return Response(
                {'error': '採購數量必須大於 0'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                org = requirement.production_order.organization
                supplier_name = requirement.supplier or 'Unknown Supplier'

                # Find or create supplier
                supplier, _ = Supplier.objects.get_or_create(
                    organization=org,
                    name=supplier_name,
                    defaults={
                        'supplier_code': f'SUP-{supplier_name[:10].upper()}',
                        'supplier_type': 'material',
                    }
                )

                # Generate PO number
                now = timezone.now()
                yymm = now.strftime('%y%m')
                count = PurchaseOrder.objects.filter(
                    organization=org,
                    po_number__startswith=f'PO-{yymm}-'
                ).count()
                po_number = f'PO-{yymm}-{str(count + 1).zfill(4)}'

                # Create PO
                po = PurchaseOrder.objects.create(
                    organization=org,
                    po_number=po_number,
                    supplier=supplier,
                    po_type='production',
                    status='draft',
                    po_date=now.date(),
                    expected_delivery=requirement.production_order.delivery_date,
                    notes=f'From Production Order {requirement.production_order.order_number}\n'
                          f'Material: {requirement.material_name}',
                )

                # Find or create material
                material, _ = Material.objects.get_or_create(
                    organization=org,
                    article_no=requirement.supplier_article_no or f'MAT-{requirement.bom_item_id.hex[:8]}',
                    defaults={
                        'name': requirement.material_name,
                        'name_zh': requirement.material_name_zh,
                        'category': requirement.category or 'other',
                        'unit': requirement.unit,
                        'supplier': supplier,
                    }
                )

                # Get unit price
                unit_price = requirement.reviewed_unit_price or Decimal('0')
                line_amount = order_qty * unit_price

                # Create single POLine with delivery tracking
                line = POLine.objects.create(
                    purchase_order=po,
                    material=material,
                    material_name=requirement.material_name,
                    color=getattr(requirement.bom_item, 'color', '') or '',
                    quantity=order_qty,
                    unit=requirement.unit,
                    unit_price=unit_price,
                    line_total=line_amount,
                    is_confirmed=True,  # Auto-confirm since already reviewed
                    confirmed_at=now,
                    notes=requirement.review_notes,
                    # Delivery tracking
                    required_date=requirement.required_date,
                    expected_delivery=requirement.expected_delivery or requirement.production_order.delivery_date,
                    delivery_status='pending',
                )

                # Update PO total
                po.total_amount = line_amount
                po.save(update_fields=['total_amount'])

                # Update requirement status
                requirement.status = 'ordered'
                requirement.purchase_order_line = line
                requirement.save()

                return Response({
                    'id': str(requirement.id),
                    'status': 'ordered',
                    'purchase_order': {
                        'id': str(po.id),
                        'po_number': po.po_number,
                        'supplier': supplier.name,
                        'quantity': float(order_qty),
                        'unit_price': float(unit_price),
                        'total_amount': float(line_amount),
                    },
                    'message': f'採購單 {po.po_number} 已生成'
                })

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
