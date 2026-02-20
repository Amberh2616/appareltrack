"""
P2: Excel Export Service for Sample Documents

Exports MWO, Estimate, and T2 PO to Excel format using openpyxl.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import json
from decimal import Decimal


class ExcelExporter:
    """Base class for Excel export with common styling utilities"""

    # Color scheme
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    LABEL_FONT = Font(bold=True, size=10)
    NORMAL_FONT = Font(size=10)

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def create_response(workbook, filename):
        """Create HTTP response with Excel file"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    @staticmethod
    def auto_width(ws, max_width=80):
        """Auto-adjust column widths"""
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, max_width)

    @staticmethod
    def apply_header_style(cell):
        """Apply header style to a cell"""
        cell.font = ExcelExporter.HEADER_FONT
        cell.fill = ExcelExporter.HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = ExcelExporter.THIN_BORDER

    @staticmethod
    def apply_subheader_style(cell):
        """Apply subheader style to a cell"""
        cell.font = ExcelExporter.LABEL_FONT
        cell.fill = ExcelExporter.SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = ExcelExporter.THIN_BORDER

    @staticmethod
    def format_decimal(value):
        """Format decimal values"""
        if value is None:
            return ''
        if isinstance(value, (int, float, Decimal)):
            return f"{float(value):.2f}"
        return str(value)


class MWOExcelExporter(ExcelExporter):
    """Export SampleMWO to Excel"""

    def export(self, mwo):
        wb = Workbook()

        # Sheet 1: Overview
        ws_overview = wb.active
        ws_overview.title = "MWO Overview"
        self._write_overview(ws_overview, mwo)

        # Sheet 2: BOM
        ws_bom = wb.create_sheet("BOM")
        self._write_bom(ws_bom, mwo)

        # Sheet 3: Operations
        ws_ops = wb.create_sheet("Operations")
        self._write_operations(ws_ops, mwo)

        # Sheet 4: QC Checkpoints
        ws_qc = wb.create_sheet("QC")
        self._write_qc(ws_qc, mwo)

        # Sheet 5: Spec (Measurements)
        ws_spec = wb.create_sheet("Spec")
        self._write_spec(ws_spec, mwo)

        # Auto-adjust widths
        for ws in wb.worksheets:
            self.auto_width(ws)

        filename = f"MWO_{mwo.mwo_no}.xlsx"
        return self.create_response(wb, filename)

    def _write_overview(self, ws, mwo):
        """Write MWO header information"""
        # Title
        ws['A1'] = 'MANUFACTURING WORK ORDER'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        row = 3
        # Basic Info
        data = [
            ('MWO Number:', mwo.mwo_no or 'N/A'),
            ('Factory:', mwo.factory_name or 'N/A'),
            ('Status:', mwo.get_status_display() if hasattr(mwo, 'get_status_display') else str(mwo.status)),
            ('Version:', f"v{mwo.version_no}" if hasattr(mwo, 'version_no') else 'v1'),
            ('Start Date:', mwo.start_date.strftime('%Y-%m-%d') if mwo.start_date else 'N/A'),
            ('Due Date:', mwo.due_date.strftime('%Y-%m-%d') if mwo.due_date else 'N/A'),
            ('Quantity:', getattr(mwo, 'quantity', 0) or 0),
            ('Created:', mwo.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(mwo, 'created_at') and mwo.created_at else 'N/A'),
        ]

        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.LABEL_FONT
            ws[f'B{row}'] = str(value)
            row += 1

        # Notes
        if hasattr(mwo, 'notes') and mwo.notes:
            row += 1
            ws[f'A{row}'] = 'Notes:'
            ws[f'A{row}'].font = self.LABEL_FONT
            row += 1
            ws[f'A{row}'] = mwo.notes
            ws.merge_cells(f'A{row}:B{row}')

    def _write_bom(self, ws, mwo):
        """Write BOM snapshot to sheet"""
        # Parse BOM snapshot JSON
        bom_data = getattr(mwo, 'bom_snapshot_json', None) or []

        # Fallback: If snapshot is empty, read from guidance_usage
        if not bom_data:
            try:
                run = mwo.sample_run
                if hasattr(run, 'guidance_usage') and run.guidance_usage:
                    usage_lines = run.guidance_usage.usage_lines.select_related('bom_item').all()
                    bom_data = []
                    for idx, ul in enumerate(usage_lines, 1):
                        bom_item = ul.bom_item
                        bom_data.append({
                            'line_no': idx,
                            'material_name': bom_item.material_name,
                            'material_name_zh': getattr(bom_item, 'material_name_zh', ''),
                            'color': getattr(bom_item, 'color', ''),
                            'uom': ul.consumption_unit or getattr(bom_item, 'uom', ''),
                            'consumption': float(ul.consumption) if ul.consumption else 0,
                            'wastage_pct': float(getattr(bom_item, 'wastage_pct', 0) or 0),
                            'unit_price': float(getattr(bom_item, 'unit_price', 0) or 0),
                            'supplier_name': getattr(bom_item, 'supplier', '') or '',
                        })
            except Exception as e:
                # If fallback fails, show error message
                ws['A1'] = f'Error loading BOM data: {str(e)}'
                return

        if not bom_data or not isinstance(bom_data, list):
            ws['A1'] = 'No BOM data available'
            return

        # Headers
        headers = ['Line', 'Material', 'Material (ZH)', 'Color', 'UOM', 'Consumption', 'Wastage %', 'Unit Price', 'Supplier']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.apply_header_style(cell)

        # Data rows
        for row_idx, item in enumerate(bom_data, 2):
            ws.cell(row=row_idx, column=1, value=item.get('line_no', row_idx - 1))
            ws.cell(row=row_idx, column=2, value=item.get('material_name', ''))
            ws.cell(row=row_idx, column=3, value=item.get('material_name_zh', ''))
            ws.cell(row=row_idx, column=4, value=item.get('color', ''))
            ws.cell(row=row_idx, column=5, value=item.get('uom', ''))
            ws.cell(row=row_idx, column=6, value=self.format_decimal(item.get('consumption', 0)))
            ws.cell(row=row_idx, column=7, value=self.format_decimal(item.get('wastage_pct', 0)))
            ws.cell(row=row_idx, column=8, value=self.format_decimal(item.get('unit_price', 0)))
            ws.cell(row=row_idx, column=9, value=item.get('supplier_name', ''))

            # Apply borders
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).border = self.THIN_BORDER

    def _write_operations(self, ws, mwo):
        """Write operations/construction snapshot"""
        ops_data = getattr(mwo, 'construction_snapshot_json', None) or []

        if not ops_data or not isinstance(ops_data, list):
            ws['A1'] = 'No operations data available'
            return

        # Headers
        headers = ['Step', 'Operation', 'Description', 'Time (min)', 'Machine']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.apply_header_style(cell)

        # Data rows
        for row_idx, item in enumerate(ops_data, 2):
            ws.cell(row=row_idx, column=1, value=item.get('step_no', row_idx - 1))
            ws.cell(row=row_idx, column=2, value=item.get('operation', ''))
            ws.cell(row=row_idx, column=3, value=item.get('description', ''))
            ws.cell(row=row_idx, column=4, value=self.format_decimal(item.get('time_minutes', 0)))
            ws.cell(row=row_idx, column=5, value=item.get('machine_type', ''))

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = self.THIN_BORDER

    def _write_qc(self, ws, mwo):
        """Write QC checkpoints"""
        qc_data = getattr(mwo, 'qc_snapshot_json', None) or []

        if not qc_data or not isinstance(qc_data, list):
            ws['A1'] = 'No QC data available'
            return

        # Headers
        headers = ['#', 'Checkpoint', 'Description', 'Standard', 'Measurement Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.apply_header_style(cell)

        # Data rows
        for row_idx, item in enumerate(qc_data, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=item.get('checkpoint_name', ''))
            ws.cell(row=row_idx, column=3, value=item.get('description', ''))
            ws.cell(row=row_idx, column=4, value=item.get('standard', ''))
            ws.cell(row=row_idx, column=5, value=item.get('measurement_type', ''))

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = self.THIN_BORDER




    def _write_spec(self, ws, mwo):
        """Write Spec/Measurement data - expands values JSON into size columns"""
        # Try to get from snapshot first
        spec_data = getattr(mwo, 'measurement_snapshot_json', None) or []
        
        # Fallback: read from StyleRevision measurements
        if not spec_data:
            try:
                run = mwo.sample_run
                if hasattr(run, 'style_revision') and run.style_revision:
                    measurements = run.style_revision.measurements.all()
                    spec_data = []
                    for m in measurements:
                        spec_data.append({
                            'point_name': m.point_name,
                            'point_name_zh': getattr(m, 'point_name_zh', ''),
                            'unit': m.unit,
                            'tol_minus': float(getattr(m, 'tolerance_minus', 0) or 0),
                            'tol_plus': float(getattr(m, 'tolerance_plus', 0) or 0),
                            'values': m.values if isinstance(m.values, dict) else {},
                        })
            except Exception as e:
                ws['A1'] = f'Error loading Spec data: {str(e)}'
                return
        
        if not spec_data or not isinstance(spec_data, list):
            ws['A1'] = 'No Spec data available'
            return
        
        # Collect all size keys from all measurements
        all_sizes = set()
        for item in spec_data:
            values = item.get('values', {})
            if isinstance(values, dict):
                all_sizes.update(values.keys())
        
        # Sort sizes: try numeric first, then string
        def sort_key(x):
            try:
                return (0, int(x))
            except ValueError:
                return (1, x)
        
        sizes = sorted(all_sizes, key=sort_key)
        
        if not sizes:
            sizes = ['0', '2', '4', '6', '8', '10', '12', '14', '16', '18']
        
        # Headers: Point Name (EN) | Point Name (ZH) | Unit | Tol- | Tol+ | Size0 | Size2 | ...
        headers = ['Point Name (EN)', 'Point Name (ZH)', 'Unit', 'Tol-', 'Tol+'] + sizes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.apply_header_style(cell)
        
        # Data rows
        for row_idx, item in enumerate(spec_data, 2):
            ws.cell(row=row_idx, column=1, value=item.get('point_name', ''))
            ws.cell(row=row_idx, column=2, value=item.get('point_name_zh', ''))
            ws.cell(row=row_idx, column=3, value=item.get('unit', ''))
            ws.cell(row=row_idx, column=4, value=self.format_decimal(item.get('tol_minus', 0)))
            ws.cell(row=row_idx, column=5, value=self.format_decimal(item.get('tol_plus', 0)))
            
            # Size values
            values = item.get('values', {})
            if isinstance(values, dict):
                for col_offset, size in enumerate(sizes):
                    val = values.get(size, values.get(str(size), ''))
                    ws.cell(row=row_idx, column=6 + col_offset, value=self.format_decimal(val) if val else '')
            
            # Apply borders
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = self.THIN_BORDER

class EstimateExcelExporter(ExcelExporter):
    """Export SampleCostEstimate to Excel"""

    def export(self, estimate):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cost Estimate"

        self._write_estimate(ws, estimate)
        self.auto_width(ws)

        filename = f"Estimate_{estimate.id}.xlsx"
        return self.create_response(wb, filename)

    def _write_estimate(self, ws, estimate):
        """Write estimate breakdown"""
        # Title
        ws['A1'] = 'SAMPLE COST ESTIMATE'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        row = 3
        # Basic Info
        data = [
            ('Estimate Version:', f"v{getattr(estimate, 'estimate_version', 1)}"),
            ('Status:', estimate.get_status_display() if hasattr(estimate, 'get_status_display') else str(estimate.status)),
            ('Currency:', getattr(estimate, 'currency', 'USD')),
            ('Valid Until:', estimate.valid_until.strftime('%Y-%m-%d') if estimate.valid_until else 'N/A'),
            ('Created:', estimate.created_at.strftime('%Y-%m-%d') if hasattr(estimate, 'created_at') and estimate.created_at else 'N/A'),
        ]

        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.LABEL_FONT
            ws[f'B{row}'] = str(value)
            row += 1

        row += 1
        ws[f'A{row}'] = 'COST BREAKDOWN'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Parse breakdown JSON
        breakdown = getattr(estimate, 'breakdown_snapshot_json', None) or {}

        # Materials
        if 'materials' in breakdown:
            row += 1
            ws[f'A{row}'] = 'Materials:'
            self.apply_subheader_style(ws[f'A{row}'])
            ws[f'B{row}'] = self.format_decimal(breakdown['materials'])
            ws[f'B{row}'].font = self.NORMAL_FONT
            ws[f'B{row}'].alignment = Alignment(horizontal='right')

        # Labor
        if 'labor' in breakdown:
            row += 1
            ws[f'A{row}'] = 'Labor:'
            self.apply_subheader_style(ws[f'A{row}'])
            ws[f'B{row}'] = self.format_decimal(breakdown['labor'])
            ws[f'B{row}'].font = self.NORMAL_FONT
            ws[f'B{row}'].alignment = Alignment(horizontal='right')

        # Overhead
        if 'overhead' in breakdown:
            row += 1
            ws[f'A{row}'] = 'Overhead:'
            self.apply_subheader_style(ws[f'A{row}'])
            ws[f'B{row}'] = self.format_decimal(breakdown['overhead'])
            ws[f'B{row}'].font = self.NORMAL_FONT
            ws[f'B{row}'].alignment = Alignment(horizontal='right')

        # Subtotal
        row += 1
        ws[f'A{row}'] = 'Subtotal:'
        ws[f'A{row}'].font = self.LABEL_FONT
        subtotal = breakdown.get('materials', 0) + breakdown.get('labor', 0) + breakdown.get('overhead', 0)
        ws[f'B{row}'] = self.format_decimal(subtotal)
        ws[f'B{row}'].font = self.LABEL_FONT
        ws[f'B{row}'].alignment = Alignment(horizontal='right')

        # Margin
        if 'margin' in breakdown:
            row += 1
            ws[f'A{row}'] = f"Margin ({breakdown.get('margin_pct', 0)}%):"
            ws[f'A{row}'].font = self.NORMAL_FONT
            ws[f'B{row}'] = self.format_decimal(breakdown['margin'])
            ws[f'B{row}'].alignment = Alignment(horizontal='right')

        # Total
        row += 1
        ws[f'A{row}'] = 'TOTAL:'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = self.format_decimal(estimate.estimated_total)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].alignment = Alignment(horizontal='right')

        # Borders for summary section
        for r in range(row - 4, row + 1):
            for c in range(1, 3):
                ws.cell(row=r, column=c).border = self.THIN_BORDER


class T2POExcelExporter(ExcelExporter):
    """Export T2POForSample to Excel"""

    def export(self, po):
        wb = Workbook()

        # Sheet 1: PO Header
        ws_header = wb.active
        ws_header.title = "PO Header"
        self._write_header(ws_header, po)

        # Sheet 2: Line Items
        ws_lines = wb.create_sheet("Line Items")
        self._write_lines(ws_lines, po)

        # Auto-adjust widths
        for ws in wb.worksheets:
            self.auto_width(ws)

        filename = f"T2PO_{po.po_no}.xlsx"
        return self.create_response(wb, filename)

    def _write_header(self, ws, po):
        """Write PO header info"""
        # Title
        ws['A1'] = 'TIER 2 PURCHASE ORDER (SAMPLE)'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        row = 3
        # Basic Info
        data = [
            ('PO Number:', po.po_no or 'N/A'),
            ('Supplier:', po.supplier_name or 'N/A'),
            ('Status:', po.get_status_display() if hasattr(po, 'get_status_display') else str(po.status)),
            ('Version:', f"v{getattr(po, 'version_no', 1)}"),
            ('Issue Date:', po.issue_date.strftime('%Y-%m-%d') if getattr(po, 'issue_date', None) else 'N/A'),
            ('Delivery Date:', po.delivery_date.strftime('%Y-%m-%d') if po.delivery_date else 'N/A'),
            ('Total Amount:', self.format_decimal(getattr(po, 'total_amount', 0))),
            ('Currency:', getattr(po, 'currency', 'USD')),
            ('Created:', po.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(po, 'created_at') and po.created_at else 'N/A'),
        ]

        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.LABEL_FONT
            ws[f'B{row}'] = str(value)
            row += 1

        # Notes
        if hasattr(po, 'notes') and po.notes:
            row += 1
            ws[f'A{row}'] = 'Notes:'
            ws[f'A{row}'].font = self.LABEL_FONT
            row += 1
            ws[f'A{row}'] = po.notes
            ws.merge_cells(f'A{row}:B{row}')

    def _write_lines(self, ws, po):
        """Write PO line items"""
        # Query line items
        try:
            lines = list(po.lines.all().order_by('line_no'))
        except Exception as e:
            ws['A1'] = f'Error loading line items: {str(e)}'
            return

        if not lines:
            ws['A1'] = 'No line items available'
            return

        # Headers
        headers = ['Line', 'Material', 'Material (ZH)', 'Color', 'UOM', 'Quantity', 'Wastage', 'Unit Price', 'Line Total', 'Supplier']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.apply_header_style(cell)

        # Data rows
        for row_idx, line in enumerate(lines, 2):
            ws.cell(row=row_idx, column=1, value=getattr(line, 'line_no', row_idx - 1))
            ws.cell(row=row_idx, column=2, value=getattr(line, 'material_name', '') or '')
            ws.cell(row=row_idx, column=3, value=getattr(line, 'material_name_zh', '') or '')
            ws.cell(row=row_idx, column=4, value=getattr(line, 'color', '') or '')
            ws.cell(row=row_idx, column=5, value=getattr(line, 'uom', '') or '')
            ws.cell(row=row_idx, column=6, value=self.format_decimal(getattr(line, 'quantity', 0)))
            ws.cell(row=row_idx, column=7, value=self.format_decimal(getattr(line, 'wastage_quantity', 0)))
            ws.cell(row=row_idx, column=8, value=self.format_decimal(getattr(line, 'unit_price', 0)))
            ws.cell(row=row_idx, column=9, value=self.format_decimal(getattr(line, 'line_total', 0)))
            ws.cell(row=row_idx, column=10, value=getattr(line, 'supplier_name', '') or '')

            # Apply borders
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).border = self.THIN_BORDER

        # Total row
        total_row = len(lines) + 2
        ws.cell(row=total_row, column=8, value='TOTAL:')
        ws.cell(row=total_row, column=8).font = self.LABEL_FONT
        ws.cell(row=total_row, column=8).alignment = Alignment(horizontal='right')

        total_amount = sum(getattr(line, 'line_total', 0) or 0 for line in lines)
        ws.cell(row=total_row, column=9, value=self.format_decimal(total_amount))
        ws.cell(row=total_row, column=9).font = self.LABEL_FONT
        ws.cell(row=total_row, column=9).border = self.THIN_BORDER
